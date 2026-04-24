"""Generate Giordano's DTC daily revenue report.

The script pulls daily performance from Shopify, Klaviyo, and Meta Ads,
normalizes it into a single table, and writes reports/daily_revenue.xlsx.
Credentials are read only from environment variables.
"""

from __future__ import annotations

import base64
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
REPORT_PATH = ROOT / "reports" / "daily_revenue.xlsx"
DATA_DIR = ROOT / "data"
LOCAL_TZ = ZoneInfo(os.getenv("REPORT_TIMEZONE", "America/Chicago"))

DEFAULT_SHOPIFY_STORE = "giordanos-frozen-pizza.myshopify.com"
DEFAULT_META_AD_ACCOUNT = "act_1271027757410529"
DEFAULT_KLAVIYO_REVISION = "2025-10-15"
DEFAULT_SHOPIFY_API_VERSION = "2026-01"
DEFAULT_REPORT_START_DATE = "2026-03-03"
DEFAULT_DAILY_TARGET = Decimal("30000.00")


@dataclass(frozen=True)
class Config:
    shopify_store: str
    shopify_api_key: str
    shopify_password: str
    klaviyo_api_key: str
    meta_access_token: str
    meta_ad_account: str
    klaviyo_revision: str
    shopify_api_version: str
    start_date: date
    end_date: date
    daily_revenue_target: Decimal
    klaviyo_conversion_metric_id: str | None


def parse_date(value: str) -> date:
    return date.fromisoformat(value)


def yesterday_local() -> date:
    return datetime.now(LOCAL_TZ).date() - timedelta(days=1)


def env_value(name: str, default: str | None = None) -> str | None:
    value = os.getenv(name)
    return value if value not in (None, "") else default


def load_config() -> Config:
    missing = [
        name
        for name in ["SHOPIFY_API_KEY", "SHOPIFY_PASSWORD", "KLAVIYO_API_KEY", "META_ACCESS_TOKEN"]
        if not os.getenv(name)
    ]
    if missing:
        raise RuntimeError(f"Missing required environment variables: {', '.join(missing)}")

    start_date = parse_date(env_value("REPORT_START_DATE", DEFAULT_REPORT_START_DATE))
    end_date = parse_date(env_value("REPORT_END_DATE", yesterday_local().isoformat()))
    if end_date < start_date:
        raise RuntimeError("REPORT_END_DATE must be on or after REPORT_START_DATE")

    return Config(
        shopify_store=env_value("SHOPIFY_STORE", DEFAULT_SHOPIFY_STORE),
        shopify_api_key=os.environ["SHOPIFY_API_KEY"],
        shopify_password=os.environ["SHOPIFY_PASSWORD"],
        klaviyo_api_key=os.environ["KLAVIYO_API_KEY"],
        meta_access_token=os.environ["META_ACCESS_TOKEN"],
        meta_ad_account=env_value("META_AD_ACCOUNT", DEFAULT_META_AD_ACCOUNT),
        klaviyo_revision=env_value("KLAVIYO_REVISION", DEFAULT_KLAVIYO_REVISION),
        shopify_api_version=env_value("SHOPIFY_API_VERSION", DEFAULT_SHOPIFY_API_VERSION),
        start_date=start_date,
        end_date=end_date,
        daily_revenue_target=Decimal(env_value("DAILY_REVENUE_TARGET", str(DEFAULT_DAILY_TARGET))),
        klaviyo_conversion_metric_id=env_value("KLAVIYO_CONVERSION_METRIC_ID"),
    )


def daterange(start: date, end: date) -> list[date]:
    days: list[date] = []
    current = start
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def money(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value))


def as_float(value: Decimal | int | float) -> float:
    return float(value)


def request_json(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    params: dict[str, str] | None = None,
    method: str = "GET",
    body: dict[str, Any] | None = None,
    timeout: int = 60,
    retries: int = 3,
) -> dict[str, Any]:
    if params:
        url = f"{url}{'&' if '?' in url else '?'}{urllib.parse.urlencode(params)}"
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req_headers = dict(headers or {})
    if body is not None:
        req_headers.setdefault("Content-Type", "application/json")

    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, data=data, headers=req_headers, method=method)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode("utf-8")
                return json.loads(raw) if raw else {}
        except urllib.error.HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            if exc.code in {429, 500, 502, 503, 504} and attempt < retries - 1:
                time.sleep(2**attempt)
                continue
            raise RuntimeError(f"API request failed ({exc.code}) for {url}: {raw[:1000]}") from exc
        except urllib.error.URLError as exc:
            if attempt < retries - 1:
                time.sleep(2**attempt)
                continue
            raise RuntimeError(f"API request failed for {url}: {exc}") from exc
    raise RuntimeError(f"API request failed for {url}")


def shopify_headers(config: Config) -> dict[str, str]:
    headers = {"Accept": "application/json"}
    # Modern Shopify custom apps use an access token in the password slot.
    if config.shopify_password.startswith("shpat_"):
        headers["X-Shopify-Access-Token"] = config.shopify_password
    else:
        token = base64.b64encode(f"{config.shopify_api_key}:{config.shopify_password}".encode()).decode()
        headers["Authorization"] = f"Basic {token}"
    return headers


def fetch_shopify(config: Config) -> dict[date, dict[str, Decimal]]:
    """Fetch daily Online Store revenue, orders, and AOV from ShopifyQL."""
    query = """
    query($q: String!) {
      shopifyqlQuery(query: $q) {
        tableData {
          columns { name dataType displayName }
          rows
        }
        parseErrors
      }
    }
    """
    shopifyql = (
        "FROM sales SHOW total_sales, orders, average_order_value "
        "WHERE sales_channel = 'Online Store' "
        f"SINCE {config.start_date.isoformat()} UNTIL {config.end_date.isoformat()} "
        f"TIMESERIES day WITH TIMEZONE '{LOCAL_TZ.key}'"
    )
    data = request_json(
        f"https://{config.shopify_store}/admin/api/{config.shopify_api_version}/graphql.json",
        headers=shopify_headers(config),
        method="POST",
        body={"query": query, "variables": {"q": shopifyql}},
    )
    result = data["data"]["shopifyqlQuery"]
    if result.get("parseErrors"):
        raise RuntimeError(f"ShopifyQL parse errors: {result['parseErrors']}")

    out = {
        day: {"shopify_revenue": Decimal("0"), "orders": Decimal("0"), "aov": Decimal("0")}
        for day in daterange(config.start_date, config.end_date)
    }
    for row in result["tableData"]["rows"]:
        day = parse_date(row["day"])
        orders = money(row.get("orders"))
        revenue = money(row.get("total_sales"))
        aov = money(row.get("average_order_value")) if row.get("average_order_value") else (
            revenue / orders if orders else Decimal("0")
        )
        out[day] = {"shopify_revenue": revenue, "orders": orders, "aov": aov}
    return out


def klaviyo_headers(config: Config) -> dict[str, str]:
    return {
        "Authorization": f"Klaviyo-API-Key {config.klaviyo_api_key}",
        "Accept": "application/json",
        "revision": config.klaviyo_revision,
    }


def get_klaviyo_conversion_metric_id(config: Config) -> str:
    if config.klaviyo_conversion_metric_id:
        return config.klaviyo_conversion_metric_id

    data = request_json("https://a.klaviyo.com/api/metrics/", headers=klaviyo_headers(config))
    candidates: list[tuple[str, str]] = []
    for metric in data.get("data", []):
        attrs = metric.get("attributes", {})
        name = attrs.get("name", "")
        integration = (attrs.get("integration") or {}).get("name", "")
        if name == "Placed Order" and integration == "Shopify":
            return metric["id"]
        if name == "Placed Order":
            candidates.append((metric["id"], integration))
    if candidates:
        return candidates[0][0]
    raise RuntimeError("Could not find a Klaviyo Placed Order metric. Set KLAVIYO_CONVERSION_METRIC_ID.")


def report_timeframe(config: Config) -> dict[str, str]:
    start = datetime(config.start_date.year, config.start_date.month, config.start_date.day, tzinfo=LOCAL_TZ)
    end_next = datetime(config.end_date.year, config.end_date.month, config.end_date.day, tzinfo=LOCAL_TZ) + timedelta(
        days=1
    )
    return {"start": start.isoformat(), "end": end_next.isoformat()}


def campaign_day_from_name(name: str, year: int) -> date | None:
    match = re.search(r"\((\d{1,2})/(\d{1,2})\)\s*$", name or "")
    if not match:
        return None
    return date(year, int(match.group(1)), int(match.group(2)))


def fetch_klaviyo_campaign_revenue(config: Config, metric_id: str) -> dict[date, dict[str, Decimal]]:
    body = {
        "data": {
            "type": "campaign-values-report",
            "attributes": {
                "timeframe": report_timeframe(config),
                "conversion_metric_id": metric_id,
                "statistics": ["conversion_value"],
                "group_by": ["campaign_message_id", "campaign_id", "campaign_message_name", "send_channel"],
            },
        }
    }
    data = request_json(
        "https://a.klaviyo.com/api/campaign-values-reports/",
        headers=klaviyo_headers(config),
        method="POST",
        body=body,
    )
    out: dict[date, dict[str, Decimal]] = defaultdict(lambda: {"email": Decimal("0"), "sms": Decimal("0")})
    campaign_dates: dict[str, date | None] = {}

    for result in data["data"]["attributes"].get("results", []):
        group = result.get("groupings", {})
        campaign_id = group.get("campaign_id")
        channel = (group.get("send_channel") or "").lower()
        if channel not in {"email", "sms"} or not campaign_id:
            continue

        if campaign_id not in campaign_dates:
            parsed_day = campaign_day_from_name(group.get("campaign_message_name", ""), config.start_date.year)
            if parsed_day and config.start_date <= parsed_day <= config.end_date:
                campaign_dates[campaign_id] = parsed_day
            else:
                meta = request_json(f"https://a.klaviyo.com/api/campaigns/{campaign_id}/", headers=klaviyo_headers(config))
                attrs = meta["data"]["attributes"]
                send_time = attrs.get("send_time") or attrs.get("scheduled_at") or attrs.get("created_at")
                campaign_dates[campaign_id] = (
                    datetime.fromisoformat(send_time.replace("Z", "+00:00")).astimezone(LOCAL_TZ).date()
                    if send_time
                    else None
                )
                time.sleep(0.05)

        day = campaign_dates[campaign_id]
        if day and config.start_date <= day <= config.end_date:
            out[day][channel] += money(result.get("statistics", {}).get("conversion_value"))
    return out


def fetch_klaviyo_flow_revenue(config: Config, metric_id: str) -> dict[date, dict[str, Decimal]]:
    body = {
        "data": {
            "type": "flow-series-report",
            "attributes": {
                "timeframe": report_timeframe(config),
                "interval": "daily",
                "conversion_metric_id": metric_id,
                "statistics": ["conversion_value"],
                "group_by": ["flow_message_id", "flow_id", "flow_name", "send_channel"],
            },
        }
    }
    data = request_json(
        "https://a.klaviyo.com/api/flow-series-reports/",
        headers=klaviyo_headers(config),
        method="POST",
        body=body,
    )
    attrs = data["data"]["attributes"]
    dates = [parse_date(value[:10]) for value in attrs["date_times"]]
    out: dict[date, dict[str, Decimal]] = defaultdict(lambda: {"email": Decimal("0"), "sms": Decimal("0")})

    for result in attrs.get("results", []):
        channel = (result.get("groupings", {}).get("send_channel") or "").lower()
        if channel not in {"email", "sms"}:
            continue
        values = result.get("statistics", {}).get("conversion_value", [])
        for idx, day in enumerate(dates):
            if config.start_date <= day <= config.end_date:
                out[day][channel] += money(values[idx] if idx < len(values) else 0)
    return out


def fetch_klaviyo(config: Config) -> dict[date, dict[str, Decimal]]:
    metric_id = get_klaviyo_conversion_metric_id(config)
    campaign = fetch_klaviyo_campaign_revenue(config, metric_id)
    flow = fetch_klaviyo_flow_revenue(config, metric_id)
    out = {day: {"email_revenue": Decimal("0"), "sms_revenue": Decimal("0")} for day in daterange(config.start_date, config.end_date)}
    for day in out:
        out[day]["email_revenue"] = campaign[day]["email"] + flow[day]["email"]
        out[day]["sms_revenue"] = campaign[day]["sms"] + flow[day]["sms"]
    return out


def fetch_meta(config: Config) -> dict[date, dict[str, Decimal]]:
    out = {
        day: {"meta_revenue": Decimal("0"), "meta_spend": Decimal("0"), "meta_purchases": Decimal("0")}
        for day in daterange(config.start_date, config.end_date)
    }
    url = f"https://graph.facebook.com/v25.0/{config.meta_ad_account}/insights"
    params = {
        "access_token": config.meta_access_token,
        "time_range": json.dumps({"since": config.start_date.isoformat(), "until": config.end_date.isoformat()}),
        "fields": "date_start,date_stop,spend,action_values,actions",
        "level": "account",
        "time_increment": "1",
        "limit": "200",
    }
    while url:
        data = request_json(url, params=params if "?" not in url else None)
        params = None
        for row in data.get("data", []):
            day = parse_date(row["date_start"])
            revenue = Decimal("0")
            purchases = Decimal("0")
            for action in row.get("action_values", []) or []:
                if action.get("action_type") == "offsite_conversion.fb_pixel_purchase":
                    revenue = money(action.get("value"))
                    break
                if action.get("action_type") == "purchase":
                    revenue = money(action.get("value"))
            for action in row.get("actions", []) or []:
                if action.get("action_type") == "offsite_conversion.fb_pixel_purchase":
                    purchases = money(action.get("value"))
                    break
                if action.get("action_type") == "purchase":
                    purchases = money(action.get("value"))
            if day in out:
                out[day] = {
                    "meta_revenue": revenue,
                    "meta_spend": money(row.get("spend")),
                    "meta_purchases": purchases,
                }
        url = data.get("paging", {}).get("next")
    return out


def build_rows(config: Config) -> list[dict[str, Decimal | str]]:
    print("Fetching Shopify...", flush=True)
    shopify = fetch_shopify(config)
    print("Fetching Klaviyo...", flush=True)
    klaviyo = fetch_klaviyo(config)
    print("Fetching Meta...", flush=True)
    meta = fetch_meta(config)

    rows: list[dict[str, Decimal | str]] = []
    for day in daterange(config.start_date, config.end_date):
        shopify_revenue = shopify[day]["shopify_revenue"]
        meta_revenue = meta[day]["meta_revenue"]
        email_revenue = klaviyo[day]["email_revenue"]
        sms_revenue = klaviyo[day]["sms_revenue"]
        owned_revenue = email_revenue + sms_revenue
        other_unattributed = shopify_revenue - meta_revenue - owned_revenue
        variance = shopify_revenue - config.daily_revenue_target
        rows.append(
            {
                "date": day.isoformat(),
                "Shopify revenue": shopify_revenue,
                "Meta revenue": meta_revenue,
                "Email revenue": email_revenue,
                "SMS revenue": sms_revenue,
                "Meta spend": meta[day]["meta_spend"],
                "Orders": shopify[day]["orders"],
                "AOV": shopify[day]["aov"],
                "Owned revenue": owned_revenue,
                "Other / unattributed revenue": other_unattributed,
                "Target revenue": config.daily_revenue_target,
                "Variance vs target": variance,
                "Meta purchases": meta[day]["meta_purchases"],
            }
        )
    return rows


def write_raw_csv(rows: list[dict[str, Decimal | str]]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    csv_path = DATA_DIR / "daily_revenue.csv"
    columns = list(rows[0].keys()) if rows else []
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        handle.write(",".join(columns) + "\n")
        for row in rows:
            values = [str(row[col]) for col in columns]
            handle.write(",".join(f'"{value}"' if "," in value else value for value in values) + "\n")


def write_excel(config: Config, rows: list[dict[str, Decimal | str]]) -> None:
    REPORT_PATH.parent.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Performance"
    headers = [
        "date",
        "Shopify revenue",
        "Meta revenue",
        "Email revenue",
        "SMS revenue",
        "Meta spend",
        "Orders",
        "AOV",
        "Owned revenue",
        "Other / unattributed revenue",
        "Target revenue",
        "Variance vs target",
        "Meta purchases",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([as_float(row[h]) if isinstance(row[h], Decimal) else row[h] for h in headers])

    style_sheet(ws, currency_cols=[2, 3, 4, 5, 6, 8, 9, 10, 11, 12], integer_cols=[7, 13])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"

    summary = wb.create_sheet("Summary")
    write_summary(summary, rows)

    notes = wb.create_sheet("Sources & Notes")
    write_notes(notes, config)

    wb.save(REPORT_PATH)


def style_sheet(ws, currency_cols: list[int], integer_cols: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9DEE7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
        for col in currency_cols:
            ws.cell(row=row[0].row, column=col).number_format = '$#,##0.00;[Red]($#,##0.00)'
        for col in integer_cols:
            ws.cell(row=row[0].row, column=col).number_format = "#,##0"
    widths = [14, 18, 16, 16, 14, 14, 12, 12, 16, 24, 16, 18, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_summary(ws, rows: list[dict[str, Decimal | str]]) -> None:
    totals = {
        "Shopify revenue": sum(money(row["Shopify revenue"]) for row in rows),
        "Meta revenue": sum(money(row["Meta revenue"]) for row in rows),
        "Email revenue": sum(money(row["Email revenue"]) for row in rows),
        "SMS revenue": sum(money(row["SMS revenue"]) for row in rows),
        "Owned revenue": sum(money(row["Owned revenue"]) for row in rows),
        "Meta spend": sum(money(row["Meta spend"]) for row in rows),
        "Orders": sum(money(row["Orders"]) for row in rows),
        "Other / unattributed revenue": sum(money(row["Other / unattributed revenue"]) for row in rows),
        "Target revenue": sum(money(row["Target revenue"]) for row in rows),
        "Variance vs target": sum(money(row["Variance vs target"]) for row in rows),
        "Meta purchases": sum(money(row["Meta purchases"]) for row in rows),
    }
    shopify_total = totals["Shopify revenue"]
    rows_out = [
        ("Total Shopify revenue", totals["Shopify revenue"]),
        ("Total Meta revenue", totals["Meta revenue"]),
        ("Total Email revenue", totals["Email revenue"]),
        ("Total SMS revenue", totals["SMS revenue"]),
        ("Total Owned revenue", totals["Owned revenue"]),
        ("Total Other / unattributed revenue", totals["Other / unattributed revenue"]),
        ("Total Meta spend", totals["Meta spend"]),
        ("Total Orders", totals["Orders"]),
        ("Blended AOV", totals["Shopify revenue"] / totals["Orders"] if totals["Orders"] else Decimal("0")),
        ("Meta ROAS", totals["Meta revenue"] / totals["Meta spend"] if totals["Meta spend"] else Decimal("0")),
        ("Total target revenue", totals["Target revenue"]),
        ("Variance vs target", totals["Variance vs target"]),
        ("Meta mix", totals["Meta revenue"] / shopify_total if shopify_total else Decimal("0")),
        ("Email mix", totals["Email revenue"] / shopify_total if shopify_total else Decimal("0")),
        ("SMS mix", totals["SMS revenue"] / shopify_total if shopify_total else Decimal("0")),
        ("Owned mix", totals["Owned revenue"] / shopify_total if shopify_total else Decimal("0")),
        (
            "Other / unattributed mix",
            totals["Other / unattributed revenue"] / shopify_total if shopify_total else Decimal("0"),
        ),
    ]
    ws.append(["Metric", "Value"])
    for label, value in rows_out:
        ws.append([label, as_float(value)])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    percent_labels = {"Meta mix", "Email mix", "SMS mix", "Owned mix", "Other / unattributed mix"}
    ratio_labels = {"Meta ROAS"}
    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row_idx, 1).value
        if label in percent_labels:
            ws.cell(row_idx, 2).number_format = "0.0%"
        elif label in ratio_labels:
            ws.cell(row_idx, 2).number_format = "0.00"
        elif label == "Total Orders":
            ws.cell(row_idx, 2).number_format = "#,##0"
        else:
            ws.cell(row_idx, 2).number_format = '$#,##0.00;[Red]($#,##0.00)'
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18


def write_notes(ws, config: Config) -> None:
    notes = [
        ("Generated at", datetime.now(LOCAL_TZ).isoformat(timespec="seconds")),
        ("Date range", f"{config.start_date.isoformat()} through {config.end_date.isoformat()}"),
        ("Shopify", "ShopifyQL sales report filtered to sales_channel = 'Online Store'."),
        ("Klaviyo", "Campaign Values Report plus Flow Series Report using Placed Order conversion value."),
        ("Meta", "Account-level Ads Insights. Revenue uses offsite_conversion.fb_pixel_purchase action value."),
        ("Owned revenue", "Email revenue + SMS revenue."),
        ("Other / unattributed revenue", "Shopify revenue - Meta revenue - Email revenue - SMS revenue."),
        ("Variance vs target", "Shopify revenue - DAILY_REVENUE_TARGET. Target is a placeholder env variable."),
        ("Attribution caveat", "Platform attribution can overlap, so Other / unattributed is directional."),
    ]
    ws.append(["Item", "Detail"])
    for row in notes:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110


def main() -> int:
    try:
        config = load_config()
        rows = build_rows(config)
        write_raw_csv(rows)
        write_excel(config, rows)
        print(f"Wrote {REPORT_PATH}", flush=True)
        print(f"Wrote {DATA_DIR / 'daily_revenue.csv'}", flush=True)
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
